from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import (
    get_column_letter
)

from io import BytesIO

# ==========================================
# COLORES CORPORATIVOS
# ==========================================

COLOR_PRIMARIO = "0F4C81"

COLOR_SECUNDARIO = "D9EAF7"

COLOR_TOTAL = "E2F0D9"

COLOR_BORDE = "D0D7DE"

# ==========================================
# ESTILOS
# ==========================================

FUENTE_TITULO = Font(
    bold=True,
    color="FFFFFF",
    size=16
)

FUENTE_HEADER = Font(
    bold=True,
    color="FFFFFF",
    size=11
)

FUENTE_NORMAL = Font(
    size=10
)

RELLENO_TITULO = PatternFill(
    fill_type="solid",
    fgColor=COLOR_PRIMARIO
)

RELLENO_HEADER = PatternFill(
    fill_type="solid",
    fgColor=COLOR_PRIMARIO
)

RELLENO_TOTAL = PatternFill(
    fill_type="solid",
    fgColor=COLOR_TOTAL
)

BORDE = Border(

    left=Side(
        style="thin",
        color=COLOR_BORDE
    ),

    right=Side(
        style="thin",
        color=COLOR_BORDE
    ),

    top=Side(
        style="thin",
        color=COLOR_BORDE
    ),

    bottom=Side(
        style="thin",
        color=COLOR_BORDE
    )

)

CENTRO = Alignment(
    horizontal="center",
    vertical="center"
)

def crear_titulo(
    ws,
    titulo,
    columnas
):

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=columnas
    )

    cell = ws.cell(
        row=1,
        column=1
    )

    cell.value = titulo

    cell.font = FUENTE_TITULO

    cell.fill = RELLENO_TITULO

    cell.alignment = CENTRO

    ws.row_dimensions[1].height = 25


def crear_encabezados(
    ws,
    headers,
    fila=3
):

    for index, titulo in enumerate(
        headers,
        start=1
    ):

        cell = ws.cell(
            row=fila,
            column=index
        )

        cell.value = titulo

        cell.font = FUENTE_HEADER

        cell.fill = RELLENO_HEADER

        cell.border = BORDE

        cell.alignment = CENTRO


# ==========================================
# LLENAR DATOS
# ==========================================

def llenar_datos(
    ws,
    datos,
    columnas,
    fila_inicio
):
    """
    Llena una hoja de Excel a partir de una lista
    de diccionarios respetando el orden indicado
    en 'columnas'.

    Retorna la siguiente fila disponible.
    """

    fila = fila_inicio

    for registro in datos:

        for indice, campo in enumerate(
            columnas,
            start=1
        ):

            valor = registro.get(campo)

            if valor is None:
                valor = ""

            ws.cell(
                row=fila,
                column=indice
            ).value = valor

        fila += 1

    return fila


def autoajustar_columnas(
    ws,
    ancho_maximo=40
):

    for column_cells in ws.columns:

        largo = max(

            len(
                str(
                    cell.value or ""
                )
            )

            for cell in column_cells

        )

        letra = get_column_letter(
            column_cells[0].column
        )

        ws.column_dimensions[
            letra
        ].width = min(
            largo + 4,
            ancho_maximo
        )

def aplicar_filtros(
    ws,
    fila_encabezado=3
):

    ws.freeze_panes = f"A{fila_encabezado + 1}"

    ws.auto_filter.ref = ws.dimensions


# ==========================================
# CREAR HOJA REPORTE
# ==========================================

def crear_hoja_reporte(

    workbook,

    nombre_hoja,

    titulo,

    resumen,

    encabezados,

    columnas,

    datos,

    columnas_moneda=None,

    columnas_decimal=None,

    columnas_totales=None

):

    # ======================================
    # CREAR HOJA
    # ======================================

    ws = workbook.create_sheet(
        title=nombre_hoja
    )

    # ======================================
    # TITULO
    # ======================================

    crear_titulo(

        ws,

        titulo,

        len(encabezados)

    )

    # ======================================
    # RESUMEN
    # ======================================

    fila_encabezado = crear_resumen_reporte(

        ws,

        resumen

    )

    # ======================================
    # ENCABEZADOS
    # ======================================

    crear_encabezados(

        ws,

        encabezados,

        fila=fila_encabezado

    )

    # ======================================
    # DATOS
    # ======================================

    fila_fin = llenar_datos(

        ws,

        datos,

        columnas,

        fila_inicio=fila_encabezado + 1

    )

    # ======================================
    # TOTALES
    # ======================================

    if columnas_totales:

        crear_fila_totales(

            ws,

            fila=fila_fin,

            datos=datos,

            totales=columnas_totales

        )

    # ======================================
    # MONEDA
    # ======================================

    if columnas_moneda:

        aplicar_formato_moneda(

            ws,

            columnas=columnas_moneda,

            fila_inicio=fila_encabezado + 1

        )

    # ======================================
    # DECIMALES
    # ======================================

    if columnas_decimal:

        aplicar_formato_decimal(

            ws,

            columnas=columnas_decimal,

            fila_inicio=fila_encabezado + 1

        )

    # ======================================
    # BORDES
    # ======================================

    aplicar_estilo_tabla(

        ws,

        fila_inicio=fila_encabezado

    )

    # ======================================
    # AJUSTAR COLUMNAS
    # ======================================

    autoajustar_columnas(ws)

    # ======================================
    # FILTROS
    # ======================================

    aplicar_filtros(

        ws,

        fila_encabezado=fila_encabezado

    )

    return ws


# ==========================================
# GUARDAR WORKBOOK
# ==========================================

def guardar_workbook(workbook):
    """
    Guarda un Workbook en memoria y retorna
    un BytesIO listo para enviar con send_file().
    """

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return output


# ==========================================
# CREAR REPORTE EXCEL
# ==========================================
def crear_reporte_excel(
    workbook,
    hojas
):
    """
    Crea todas las hojas de un reporte Excel.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        Workbook donde se crearán las hojas.

    hojas : list[dict]
        Lista de configuraciones de hojas.
    """

    campos_requeridos = [
        "nombre_hoja",
        "titulo",
        "resumen",
        "encabezados",
        "columnas",
        "datos"
    ]

    for config in hojas:

        # ===============================
        # VALIDAR CONFIGURACIÓN
        # ===============================
        for campo in campos_requeridos:

            if campo not in config:

                raise ValueError(
                    f"Configuración inválida del reporte. "
                    f"Falta el campo obligatorio '{campo}'."
                )

        # ===============================
        # CREAR HOJA
        # ===============================
        crear_hoja_reporte(

            workbook=workbook,

            nombre_hoja=config["nombre_hoja"],

            titulo=config["titulo"],

            resumen=config["resumen"],

            encabezados=config["encabezados"],

            columnas=config["columnas"],

            datos=config["datos"],

            columnas_moneda=config.get(
                "columnas_moneda"
            ),

            columnas_decimal=config.get(
                "columnas_decimal"
            ),

            columnas_totales=config.get(
                "columnas_totales"
            )

        )

    return workbook


# ==========================================
# RESUMEN REPORTE
# ==========================================

def crear_resumen_reporte(
    ws,
    resumen,
    fila_inicio=2
):

    fila = fila_inicio

    for titulo, valor in resumen:

        ws.cell(
            row=fila,
            column=1
        ).value = titulo

        ws.cell(
            row=fila,
            column=2
        ).value = "" if valor is None else valor

        ws.cell(
            row=fila,
            column=1
        ).font = Font(
            bold=True
        )

        fila += 1

    return fila + 1


# ==========================================
# FORMATO MONEDA
# ==========================================

def aplicar_formato_moneda(
    ws,
    columnas,
    fila_inicio=2
):

    for columna in columnas:

        for row in ws.iter_rows(
            min_row=fila_inicio
        ):

            row[
                columna - 1
            ].number_format = '"COP" $ #,##0.00'


# ==========================================
# FORMATO DECIMAL
# ==========================================

def aplicar_formato_decimal(
    ws,
    columnas,
    fila_inicio=2
):

    for columna in columnas:

        for row in ws.iter_rows(
            min_row=fila_inicio
        ):

            row[
                columna - 1
            ].number_format = '#,##0.00'


# ==========================================
# TOTAL GENERAL
# ==========================================

def crear_fila_totales(
    ws,
    fila,
    datos,
    totales
):

    ws.cell(
        row=fila,
        column=1
    ).value = "TOTAL"

    ws.cell(
        row=fila,
        column=1
    ).font = Font(
        bold=True
    )

    ws.cell(
        row=fila,
        column=1
    ).fill = RELLENO_TOTAL

    for campo, columna in totales.items():

        valor = sum(

            float(
                item.get(campo) or 0
            )

            for item in datos

        )

        cell = ws.cell(
            row=fila,
            column=columna
        )

        cell.value = valor

        cell.font = Font(
            bold=True
        )

        cell.fill = RELLENO_TOTAL

        cell.number_format = '"COP" $ #,##0.00'


# ==========================================
# ESTILO TABLA
# ==========================================

def aplicar_estilo_tabla(
    ws,
    fila_inicio=1
):

    for row in ws.iter_rows(
        min_row=fila_inicio
    ):

        for cell in row:

            cell.border = BORDE